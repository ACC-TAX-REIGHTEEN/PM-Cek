import os
import re
import configparser
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIG_FILE = "config.conf"

if not os.path.exists(CONFIG_FILE):
    default_config_text = """[DEPO]
Shell KDI = KDI-0, KDI-1, KDI-4
Ban KDI = KDI-2, KDI-3
TOP 1 KDI = KDI-5
Osram KDI = KDI-6
Shell MKS = MKS-1, MKS-2, MKS-3, MKS-5
Ban MKS = MKS-4
Osram MKS = MKS-6
Shell Palu = PLU
Shell MGL = MGL
Shell SL = SL
Shell YY = YY
Shell TGL = TGL
Shell DSMG = SMG, SG
Shell PWT = PW, PWT
Shell Pati = PA

[SETTINGS]
DIGIT_DEFAULT_DEPO = Ban

[SPLIT]
Ban = Ban Jawa
MGL Shell = Selatan
SL Shell = Selatan
YY Shell = Selatan
SMG Shell = Utara
TGL Shell = Utara
PWT Shell = Utara
Pati Shell = Utara
KDI Shell = Sulawesi Shell
KDI Ban = Sulawesi Ban
KDI Top 1 = Sulawesi Top 1
KDI Osram = Sulawesi Osram
MKS Shell = Sulawesi Shell
MKS Ban = Sulawesi Ban
MKS Osram = Sulawesi Osram
Palu Shell = Sulawesi Shell
"""
    with open(CONFIG_FILE, "w") as f:
        f.write(default_config_text)
    print(f"File '{CONFIG_FILE}' belum ditemukan. File default berhasil dibuat otomatis.")

config = configparser.ConfigParser()
config.optionxform = str
config.read(CONFIG_FILE)

DIGIT_DEFAULT_DEPO = config.get("SETTINGS", "DIGIT_DEFAULT_DEPO", fallback="Ban")

PREFIX_MAP = []
ALL_DIVISIONS = [DIGIT_DEFAULT_DEPO]

if "DEPO" in config:
    for depo_name, codes_str in config["DEPO"].items():
        if depo_name not in ALL_DIVISIONS:
            ALL_DIVISIONS.append(depo_name)
            
        codes = [c.strip() for c in codes_str.split(",") if c.strip()]
        for c in codes:
            PREFIX_MAP.append((c, depo_name))

SPLIT_DICT = dict(config["SPLIT"]) if "SPLIT" in config else {}

ALL_SPLITS = []
if "SPLIT" in config:
    for val in config["SPLIT"].values():
        if val not in ALL_SPLITS:
            ALL_SPLITS.append(val)

def get_split_group(div_name):
    if div_name in SPLIT_DICT:
        return SPLIT_DICT[div_name]
    
    words_div = set(div_name.upper().replace("DSMG", "SMG").split())
    for k, v in SPLIT_DICT.items():
        words_k = set(k.upper().replace("DSMG", "SMG").split())
        if words_div == words_k:
            return v
            
    for k, v in SPLIT_DICT.items():
        if k.upper() in div_name.upper() or div_name.upper() in k.upper():
            return v
            
    return "Lainnya"

def determine_divisi(no_pel):
    val = str(no_pel).strip()
    for prefix, depo_name in PREFIX_MAP:
        if val.startswith(prefix):
            return depo_name
    if len(val) > 0 and val[0] in "123456789":
        return DIGIT_DEFAULT_DEPO
    return "Lainnya"

def determine_status_pajak(row):
    nomor_pajak = row.get("Nomor Pajak Pelanggan", "")
    nama_pelanggan = row.get("Nama Pelanggan", "")
    
    nama_str = str(nama_pelanggan).strip().upper() if pd.notna(nama_pelanggan) else ""
    if "CASH" in nama_str:
        return "Gunggung"
        
    if pd.isna(nomor_pajak):
        return "Gunggung" # Aslinya khusus untuk Gunggung Tapi Nomor Pajak Kosong
        
    val = str(nomor_pajak).strip()
    val_clean = re.sub(r'[,.]00$', '', val).replace('.', '')
    
    if val in ["", "nan", "none", "null"] or val_clean == "":
        return "Gunggung" # Aslinya khusus untuk Gunggung Tapi Nomor Pajak Kosong
        
    if val_clean in ["0", "0.0"] or (len(val_clean) > 0 and val_clean[0] in "123456789"):
        return "Tidak Gunggung"
        
    return "Gunggung (Kosong)"

def determine_cek_tanggal(tgl, tgl_pajak):
    try:
        d1 = pd.to_datetime(tgl).date()
        d2 = pd.to_datetime(tgl_pajak).date()
        return "Cocok" if d1 == d2 else "Tidak Cocok"
    except:
        return "Cocok" if str(tgl).strip() == str(tgl_pajak).strip() else "Tidak Cocok"

def clean_string_digits(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    s = re.sub(r'[,.]00$', '', s)
    s = s.replace('.', '')
    return s

input_file = "PK_temp.xlsx" if os.path.exists("PK_temp.xlsx") else "PK.xlsx"
print(f"Membaca data dari: {input_file}...")

df_raw = pd.read_excel(input_file, dtype=str)

df_raw["No. Referensi"] = df_raw["No. Referensi"].apply(clean_string_digits)
df_raw["No. Faktur Pajak"] = df_raw["No. Faktur Pajak"].apply(clean_string_digits)
df_raw["No. Pelanggan"] = df_raw["No. Pelanggan"].apply(clean_string_digits)
df_raw["Nomor Pajak Pelanggan"] = df_raw["Nomor Pajak Pelanggan"].apply(clean_string_digits)

df_raw["Jumlah Pajak"] = pd.to_numeric(
    df_raw["Jumlah Pajak"].astype(str).str.replace('.', '').str.replace(',', ''), 
    errors='coerce'
).fillna(0)

df_raw["Divisi"] = df_raw["No. Pelanggan"].apply(determine_divisi)
df_raw["Split"] = df_raw["Divisi"].apply(get_split_group)
df_raw["Status Pajak"] = df_raw.apply(determine_status_pajak, axis=1)
df_raw["Cek Tanggal"] = df_raw.apply(lambda r: determine_cek_tanggal(r["Tanggal"], r["Tgl. Pajak"]), axis=1)

wb = openpyxl.Workbook()
wb.remove(wb.active)

font_title = Font(name="Calibri", size=15, bold=True, color="1B365D")
font_subtitle = Font(name="Calibri", size=10, italic=True, color="555555")
font_section_hdr = Font(name="Calibri", size=12, bold=True, color="1B365D")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=11, bold=True)
font_regular = Font(name="Calibri", size=11)
font_alert = Font(name="Calibri", size=11, bold=True, color="9C0006")

fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
fill_teal = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
fill_summary_hdr = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
fill_light_gray = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
fill_zebra = PatternFill(start_color="F9FAFC", end_color="F9FAFC", fill_type="solid")

thin_border_side = Side(style='thin', color='D3D3D3')
thick_bottom_side = Side(style='medium', color='1B365D')
double_bottom_side = Side(style='double', color='1B365D')

border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
border_total = Border(top=thin_border_side, bottom=double_bottom_side)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

ws_dash = wb.create_sheet(title="Dashboard")
ws_dash.views.sheetView[0].showGridLines = True

ws_dash.cell(row=1, column=1, value="DASHBOARD RINGKASAN PAJAK PER WILAYAH & DIVISI").font = font_title
ws_dash.cell(row=2, column=1, value="Rangkuman Pemisahan Data & Status Pengkreditan Pajak").font = font_subtitle

ws_dash.cell(row=4, column=1, value="1. RINGKASAN TOTAL PER WILAYAH / SPLIT").font = font_section_hdr

dash_headers_split = [
    "No", "Wilayah / Split", "Pelanggan Unik", "Total Trx", 
    "Trx Gunggung", "Total Pajak Gunggung (Rp)", 
    "Trx Tidak Gunggung", "Total Pajak Tidak Gunggung (Rp)", 
    "Trx Kosong", "Total Pajak Kosong (Rp)", 
    "Total Pajak Overall (Rp)", "Kesesuaian Tanggal"
]

split_start_row = 5
for col_idx, header in enumerate(dash_headers_split, 1):
    cell = ws_dash.cell(row=split_start_row, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = fill_teal
    cell.alignment = align_center
    cell.border = border_header

current_dash_row = split_start_row + 1

for idx, split_name in enumerate(ALL_SPLITS, 1):
    df_split = df_raw[df_raw["Split"] == split_name]
    
    unique_cust = df_split["No. Pelanggan"].nunique() if len(df_split) > 0 else 0
    total_trx = len(df_split)
    
    df_g = df_split[df_split["Status Pajak"] == "Gunggung"]
    trx_g = len(df_g)
    sum_g = df_g["Jumlah Pajak"].sum()
    
    df_tg = df_split[df_split["Status Pajak"] == "Tidak Gunggung"]
    trx_tg = len(df_tg)
    sum_tg = df_tg["Jumlah Pajak"].sum()
    
    df_k = df_split[df_split["Status Pajak"] == "Gunggung (Kosong)"]
    trx_k = len(df_k)
    sum_k = df_k["Jumlah Pajak"].sum()
    
    tot_pajak = df_split["Jumlah Pajak"].sum()
    
    cocok_cnt = len(df_split[df_split["Cek Tanggal"] == "Cocok"])
    tcocok_cnt = len(df_split[df_split["Cek Tanggal"] == "Tidak Cocok"])
    
    if total_trx == 0:
        stat_tgl = "Tidak Ada Data"
    elif tcocok_cnt == 0:
        stat_tgl = "Semua Cocok (100%)"
    else:
        stat_tgl = f"{cocok_cnt} Cocok, {tcocok_cnt} Tdk Cocok"
        
    row_data = [
        idx, split_name, unique_cust, total_trx,
        trx_g, sum_g,
        trx_tg, sum_tg,
        trx_k, sum_k,
        tot_pajak, stat_tgl
    ]
    
    fill_row = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
    
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_dash.cell(row=current_dash_row, column=c_idx, value=val)
        cell.font = font_regular
        cell.fill = fill_row
        cell.border = border_cell
        
        if c_idx in [1, 3, 4, 5, 7, 9]:
            cell.alignment = align_center
            cell.number_format = '#,##0'
        elif c_idx in [6, 8, 10, 11]:
            cell.alignment = align_right
            cell.number_format = '#,##0'
        elif c_idx == 12:
            cell.alignment = align_center
            if tcocok_cnt > 0:
                cell.font = font_alert
        else:
            cell.alignment = align_left
            
    current_dash_row += 1

ws_dash.cell(row=current_dash_row, column=2, value="TOTAL KESELURUHAN WILAYAH").font = font_bold
ws_dash.cell(row=current_dash_row, column=2).alignment = align_left

sum_cols = [3, 4, 5, 6, 7, 8, 9, 10, 11]
for c_idx in range(1, len(dash_headers_split) + 1):
    cell = ws_dash.cell(row=current_dash_row, column=c_idx)
    cell.border = border_total
    cell.font = font_bold
    if c_idx in sum_cols:
        col_let = get_column_letter(c_idx)
        start_cell = f"{col_let}{split_start_row + 1}"
        end_cell = f"{col_let}{current_dash_row - 1}"
        cell.value = f"=SUM({start_cell}:{end_cell})"
        cell.number_format = '#,##0'
        cell.alignment = align_right if c_idx in [6, 8, 10, 11] else align_center

current_dash_row += 3

ws_dash.cell(row=current_dash_row, column=1, value="2. DETAIL RINGKASAN PER DIVISI").font = font_section_hdr
current_dash_row += 1

dash_headers_div = [
    "No", "Divisi / Sheet", "Pelanggan Unik", "Total Trx", 
    "Trx Gunggung", "Total Pajak Gunggung (Rp)", 
    "Trx Tidak Gunggung", "Total Pajak Tidak Gunggung (Rp)", 
    "Trx Kosong", "Total Pajak Kosong (Rp)", 
    "Total Pajak Overall (Rp)", "Kesesuaian Tanggal"
]

div_start_row = current_dash_row
for col_idx, header in enumerate(dash_headers_div, 1):
    cell = ws_dash.cell(row=div_start_row, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_center
    cell.border = border_header

current_dash_row = div_start_row + 1

for idx, div in enumerate(ALL_DIVISIONS, 1):
    df_div = df_raw[df_raw["Divisi"] == div]
    
    unique_cust = df_div["No. Pelanggan"].nunique() if len(df_div) > 0 else 0
    total_trx = len(df_div)
    
    df_g = df_div[df_div["Status Pajak"] == "Gunggung"]
    trx_g = len(df_g)
    sum_g = df_g["Jumlah Pajak"].sum()
    
    df_tg = df_div[df_div["Status Pajak"] == "Tidak Gunggung"]
    trx_tg = len(df_tg)
    sum_tg = df_tg["Jumlah Pajak"].sum()
    
    df_k = df_div[df_div["Status Pajak"] == "Gunggung (Kosong)"]
    trx_k = len(df_k)
    sum_k = df_k["Jumlah Pajak"].sum()
    
    tot_pajak = df_div["Jumlah Pajak"].sum()
    
    cocok_cnt = len(df_div[df_div["Cek Tanggal"] == "Cocok"])
    tcocok_cnt = len(df_div[df_div["Cek Tanggal"] == "Tidak Cocok"])
    
    if total_trx == 0:
        stat_tgl = "Tidak Ada Data"
    elif tcocok_cnt == 0:
        stat_tgl = "Semua Cocok (100%)"
    else:
        stat_tgl = f"{cocok_cnt} Cocok, {tcocok_cnt} Tdk Cocok"
        
    row_data = [
        idx, div, unique_cust, total_trx,
        trx_g, sum_g,
        trx_tg, sum_tg,
        trx_k, sum_k,
        tot_pajak, stat_tgl
    ]
    
    fill_row = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
    
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_dash.cell(row=current_dash_row, column=c_idx, value=val)
        cell.font = font_regular
        cell.fill = fill_row
        cell.border = border_cell
        
        if c_idx in [1, 3, 4, 5, 7, 9]:
            cell.alignment = align_center
            cell.number_format = '#,##0'
        elif c_idx in [6, 8, 10, 11]:
            cell.alignment = align_right
            cell.number_format = '#,##0'
        elif c_idx == 12:
            cell.alignment = align_center
            if tcocok_cnt > 0:
                cell.font = font_alert
        else:
            cell.alignment = align_left
            
    current_dash_row += 1

ws_dash.cell(row=current_dash_row, column=2, value="TOTAL KESELURUHAN DIVISI").font = font_bold
ws_dash.cell(row=current_dash_row, column=2).alignment = align_left

for c_idx in range(1, len(dash_headers_div) + 1):
    cell = ws_dash.cell(row=current_dash_row, column=c_idx)
    cell.border = border_total
    cell.font = font_bold
    if c_idx in sum_cols:
        col_let = get_column_letter(c_idx)
        start_cell = f"{col_let}{div_start_row + 1}"
        end_cell = f"{col_let}{current_dash_row - 1}"
        cell.value = f"=SUM({start_cell}:{end_cell})"
        cell.number_format = '#,##0'
        cell.alignment = align_right if c_idx in [6, 8, 10, 11] else align_center

detail_columns = [
    "Tanggal", "Tgl. Pajak", "No. Referensi", "No. Faktur Pajak",
    "No. Pelanggan", "Nama Pelanggan", "Negara Pelanggan",
    "Nomor Pajak Pelanggan", "Jumlah Pajak", "Status Pajak", "Cek Tanggal"
]

for div in ALL_DIVISIONS:
    ws = wb.create_sheet(title=div)
    ws.views.sheetView[0].showGridLines = True
    
    ws.cell(row=1, column=1, value=f"DATA PAJAK - DIVISI {div.upper()}").font = font_title
    
    ws.merge_cells("A3:D3")
    summary_title = ws.cell(row=3, column=1, value="RINGKASAN DIVISI")
    summary_title.font = font_header
    summary_title.fill = fill_summary_hdr
    summary_title.alignment = align_center
    
    ws.cell(row=4, column=1, value="Kategori Status Pajak").font = font_bold
    ws.cell(row=4, column=2, value="Jumlah Transaksi").font = font_bold
    ws.cell(row=4, column=3, value="Total Jumlah Pajak (Rp)").font = font_bold
    ws.cell(row=4, column=4, value="Keterangan").font = font_bold
    
    for col_i in range(1, 5):
        cell = ws.cell(row=4, column=col_i)
        cell.fill = fill_light_gray
        cell.border = border_cell
        cell.alignment = align_center if col_i in [2, 3] else align_left

    df_div = df_raw[df_raw["Divisi"] == div]
    tot_p = df_div["No. Pelanggan"].nunique() if len(df_div) > 0 else 0
    
    s_gunggung = df_div[df_div["Status Pajak"] == "Gunggung"]
    s_tgunggung = df_div[df_div["Status Pajak"] == "Tidak Gunggung"]
    s_kosong = df_div[df_div["Status Pajak"] == "Gunggung (Kosong)"]
    
    summary_rows = [
        ("Gunggung (Transaksi CASH)", len(s_gunggung), s_gunggung["Jumlah Pajak"].sum(), "Transaksi Pelanggan CASH"),
        ("Tidak Gunggung (NPWP 1-9 / Non-CASH 0)", len(s_tgunggung), s_tgunggung["Jumlah Pajak"].sum(), "Faktur Pajak / Transaksi Non-CASH"),
        ("Gunggung (Data Kosong)", len(s_kosong), s_kosong["Jumlah Pajak"].sum(), "Masuk Gunggung (NPWP Kosong/Blank)"),
    ]
    
    for idx, (cat, cnt, amt, ket) in enumerate(summary_rows, 5):
        ws.cell(row=idx, column=1, value=cat).font = font_regular
        ws.cell(row=idx, column=2, value=cnt).font = font_regular
        ws.cell(row=idx, column=3, value=amt).font = font_regular
        ws.cell(row=idx, column=4, value=ket).font = font_subtitle
        
        ws.cell(row=idx, column=1).alignment = align_left
        ws.cell(row=idx, column=2).alignment = align_center
        ws.cell(row=idx, column=3).alignment = align_right
        ws.cell(row=idx, column=4).alignment = align_left
        
        ws.cell(row=idx, column=2).number_format = '#,##0'
        ws.cell(row=idx, column=3).number_format = '#,##0'
        
        for c_i in range(1, 5):
            ws.cell(row=idx, column=c_i).border = border_cell
            
    ws.cell(row=8, column=1, value="TOTAL").font = font_bold
    ws.cell(row=8, column=2, value=f"=SUM(B5:B7)").font = font_bold
    ws.cell(row=8, column=3, value=f"=SUM(C5:C7)").font = font_bold
    ws.cell(row=8, column=4, value=f"Total Pelanggan Unik: {tot_p}").font = font_bold
    
    ws.cell(row=8, column=2).alignment = align_center
    ws.cell(row=8, column=3).alignment = align_right
    ws.cell(row=8, column=2).number_format = '#,##0'
    ws.cell(row=8, column=3).number_format = '#,##0'
    
    for c_i in range(1, 5):
        ws.cell(row=8, column=c_i).border = border_total

    data_start_row = 10
    for col_idx, col_name in enumerate(detail_columns, 1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_header
        
    current_data_row = data_start_row + 1
    
    if len(df_div) > 0:
        for r_idx, row in df_div.iterrows():
            row_values = [
                row["Tanggal"], row["Tgl. Pajak"], row["No. Referensi"], row["No. Faktur Pajak"],
                row["No. Pelanggan"], row["Nama Pelanggan"], row["Negara Pelanggan"],
                row["Nomor Pajak Pelanggan"], row["Jumlah Pajak"], row["Status Pajak"], row["Cek Tanggal"]
            ]
            
            fill_r = fill_zebra if current_data_row % 2 == 0 else PatternFill(fill_type=None)
            
            for c_idx, val in enumerate(row_values, 1):
                if c_idx in [3, 4, 5, 8]:
                    cell = ws.cell(row=current_data_row, column=c_idx, value=str(val) if pd.notna(val) else "")
                    cell.number_format = '@'
                elif c_idx == 9:
                    cell = ws.cell(row=current_data_row, column=c_idx, value=float(val) if pd.notna(val) else 0)
                    cell.number_format = '#,##0'
                else:
                    cell = ws.cell(row=current_data_row, column=c_idx, value=val)
                    
                cell.font = font_regular
                cell.fill = fill_r
                cell.border = border_cell
                
                if c_idx in [1, 2, 5, 10, 11]:
                    cell.alignment = align_center
                elif c_idx in [3, 4, 8]:
                    cell.alignment = align_center
                elif c_idx == 9:
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
                    
                if c_idx == 11 and val == "Tidak Cocok":
                    cell.font = font_alert
                    
            current_data_row += 1
            
        ws.cell(row=current_data_row, column=1, value="TOTAL").font = font_bold
        for c_i in range(1, len(detail_columns) + 1):
            cell = ws.cell(row=current_data_row, column=c_i)
            cell.border = border_total
            cell.font = font_bold
            if c_i == 9:
                cell.value = f"=SUM(I{data_start_row + 1}:I{current_data_row - 1})"
                cell.number_format = '#,##0'
                cell.alignment = align_right
    else:
        cell = ws.cell(row=current_data_row, column=1, value="Tidak ada data untuk divisi ini.")
        cell.font = font_subtitle

for ws in wb.worksheets:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        
        if ws.title != "Dashboard":
            if col_letter == 'F':
                ws.column_dimensions[col_letter].width = 20
                continue
            elif col_letter == 'G':
                ws.column_dimensions[col_letter].width = 20
                continue
        
        max_len = 0
        for cell in col:
            if cell.row in [1, 2, 3]:
                continue
            if cell.value:
                val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
        
        adjusted_width = max(max_len + 4, 12)
        ws.column_dimensions[col_letter].width = min(adjusted_width, 35)

output_filename = "PK_Processed.xlsx"
wb.save(output_filename)
print(f"Pemrosesan selesai! File berhasil disimpan sebagai: {output_filename}")